#!/usr/bin/env python3
"""
Knowledge Base Ingestion Script

Ingests web pages and local files into SQLite FTS5 knowledge base.

Usage:
    uv run skills/knowledge-base/scripts/kb_ingest.py --all
    uv run skills/knowledge-base/scripts/kb_ingest.py --source files
    uv run skills/knowledge-base/scripts/kb_ingest.py --source web
    uv run skills/knowledge-base/scripts/kb_ingest.py --file "path/to/file.xlsx"
    uv run skills/knowledge-base/scripts/kb_ingest.py --url "https://example.com"
"""

import argparse
import json
import re
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

# Project root = where `uv run` is executed from
PROJECT_ROOT = Path.cwd()
KB_DIR = PROJECT_ROOT / ".autoservice" / "database" / "knowledge_base"
CHUNKS_DIR = KB_DIR / "chunks"
SOURCES_FILE = Path(__file__).parent.parent / "references" / "sources.json"

CHUNK_MAX_CHARS = 600
CHUNK_MIN_CHARS = 50
ROWS_PER_XLSX_CHUNK = 15
ROWS_PER_RATE_TABLE_CHUNK = 3  # Smaller chunks for country/rate tables for precise lookup

# Column name patterns that identify a rate/pricing table by country
RATE_TABLE_SIGNALS = {"country", "did", "mrc", "rate", "dial-in", "dial-out", "leg1", "leg2"}

# Country name → region code mapping for XLSX rate tables
COUNTRY_REGION_MAP = {
    "united states": "US", "usa": "US", "us": "US",
    "united kingdom": "UK", "uk": "UK", "great britain": "UK",
    "hong kong": "HK", "hk": "HK",
    "singapore": "SG", "sg": "SG",
    "japan": "JP", "jp": "JP",
    "australia": "AU", "au": "AU",
    "germany": "DE", "de": "DE",
    "france": "FR", "fr": "FR",
    "canada": "CA", "ca": "CA",
    "china": "CN", "cn": "CN", "mainland china": "CN",
    "taiwan": "TW", "tw": "TW",
    "south korea": "KR", "korea": "KR", "kr": "KR",
    "india": "IN", "in": "IN",
    "indonesia": "ID", "id": "ID",
    "malaysia": "MY", "my": "MY",
    "thailand": "TH", "th": "TH",
    "philippines": "PH", "ph": "PH",
    "vietnam": "VN", "vn": "VN",
    "brazil": "BR", "br": "BR",
    "mexico": "MX", "mx": "MX",
    "netherlands": "NL", "nl": "NL",
    "italy": "IT", "it": "IT",
    "spain": "ES", "es": "ES",
    "sweden": "SE", "se": "SE",
    "switzerland": "CH", "ch": "CH",
    "new zealand": "NZ", "nz": "NZ",
    "ireland": "IE", "ie": "IE",
    "israel": "IL", "il": "IL",
    "uae": "AE", "united arab emirates": "AE",
    "south africa": "ZA", "za": "ZA",
}


# ─── Database Setup ────────────────────────────────────────────────────────────

def init_db(db_path: Path) -> sqlite3.Connection:
    """Initialize SQLite database with FTS5."""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS kb_chunks (
            id          TEXT PRIMARY KEY,
            source_id   TEXT NOT NULL,
            source_type TEXT NOT NULL,
            source_name TEXT NOT NULL,
            source_url  TEXT,
            file_path   TEXT,
            section     TEXT,
            content     TEXT NOT NULL,
            created_at  TEXT NOT NULL
        )
    """)

    # Add new columns (safe to run multiple times)
    for col_def in [
        "domain TEXT DEFAULT ''",
        "region TEXT DEFAULT ''",
        "language TEXT DEFAULT 'en'",
        "page_number INTEGER",
    ]:
        col_name = col_def.split()[0]
        try:
            conn.execute(f"ALTER TABLE kb_chunks ADD COLUMN {col_def}")
        except sqlite3.OperationalError:
            pass  # Column already exists

    # Create indexes for new columns
    conn.execute("CREATE INDEX IF NOT EXISTS idx_kb_domain ON kb_chunks(domain)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_kb_region ON kb_chunks(region)")

    conn.execute("""
        CREATE VIRTUAL TABLE IF NOT EXISTS kb_fts USING fts5(
            chunk_id UNINDEXED,
            source_name,
            section,
            content,
            content=kb_chunks,
            content_rowid=rowid,
            tokenize="unicode61 remove_diacritics 1"
        )
    """)

    conn.execute("""
        CREATE TRIGGER IF NOT EXISTS kb_ai AFTER INSERT ON kb_chunks BEGIN
            INSERT INTO kb_fts(rowid, chunk_id, source_name, section, content)
            VALUES (new.rowid, new.id, new.source_name, new.section, new.content);
        END
    """)

    conn.execute("""
        CREATE TRIGGER IF NOT EXISTS kb_ad AFTER DELETE ON kb_chunks BEGIN
            INSERT INTO kb_fts(kb_fts, rowid, chunk_id, source_name, section, content)
            VALUES ('delete', old.rowid, old.id, old.source_name, old.section, old.content);
        END
    """)

    conn.commit()
    return conn


def clear_source(conn: sqlite3.Connection, source_id: str):
    """Remove all existing chunks for a source before re-ingesting."""
    conn.execute("DELETE FROM kb_chunks WHERE source_id = ?", (source_id,))
    conn.commit()


def save_chunk(conn: sqlite3.Connection, chunk: dict, source_dir: Path):
    """Write a chunk to the database and a JSON file for debugging."""
    conn.execute(
        """INSERT OR REPLACE INTO kb_chunks
           (id, source_id, source_type, source_name, source_url, file_path,
            section, content, created_at, domain, region, language, page_number)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            chunk["id"], chunk["source_id"], chunk["source_type"],
            chunk["source_name"], chunk.get("source_url"),
            chunk.get("file_path"), chunk.get("section", ""),
            chunk["content"], chunk["created_at"],
            chunk.get("domain", ""), chunk.get("region", ""),
            chunk.get("language", "en"), chunk.get("page_number"),
        ),
    )
    source_dir.mkdir(parents=True, exist_ok=True)
    (source_dir / f"{chunk['id']}.json").write_text(
        json.dumps(chunk, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def make_chunk_id(source_id: str, index: int) -> str:
    return f"{source_id}_{index:04d}"


def chunk_paragraphs(text: str) -> list[str]:
    """Split text into chunks by double-newline paragraphs, respecting max size."""
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    chunks, current, current_len = [], [], 0

    for para in paragraphs:
        if len(para) < CHUNK_MIN_CHARS:
            continue
        if current_len + len(para) > CHUNK_MAX_CHARS and current:
            chunks.append("\n\n".join(current))
            current, current_len = [para], len(para)
        else:
            current.append(para)
            current_len += len(para)

    if current:
        chunks.append("\n\n".join(current))
    return chunks


# ─── PDF Heading Detection ────────────────────────────────────────────────────

def _is_heading(line: str) -> bool:
    """Detect if a line is likely a heading."""
    stripped = line.strip()
    if not stripped or len(stripped) > 120:
        return False
    # ALL CAPS lines (at least 3 chars, not just numbers/symbols)
    if stripped.isupper() and len(stripped) >= 3 and re.search(r"[A-Z]{3}", stripped):
        return True
    # Numbered section headings: "1. Title", "2.1 Features", "3.1.2 Details"
    if re.match(r"^\d+(\.\d+)*\.?\s+\S", stripped):
        return True
    return False


def _is_table_line(line: str) -> bool:
    """Detect if a line is part of a table (has 2+ pipe or tab separators)."""
    return line.count("|") >= 2 or line.count("\t") >= 2


def _semantic_chunk_pages(pages: list[tuple[int, str]], max_chars: int) -> list[dict]:
    """Split PDF pages into semantically-aware chunks.

    Returns list of dicts: {"text": str, "section": str, "page_start": int, "page_end": int}
    """
    # Combine all text with page markers
    segments: list[dict] = []  # {"text", "section", "page_start", "page_end", "is_table"}

    current_section = ""
    current_lines: list[str] = []
    current_page_start = 1
    current_page_end = 1
    in_table = False

    for page_num, page_text in pages:
        lines = page_text.splitlines()
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue

            # Check for heading
            if _is_heading(stripped) and not in_table:
                # Flush current segment
                if current_lines:
                    segments.append({
                        "text": "\n".join(current_lines),
                        "section": current_section,
                        "page_start": current_page_start,
                        "page_end": current_page_end,
                        "is_table": False,
                    })
                current_section = stripped
                current_lines = []
                current_page_start = page_num
                current_page_end = page_num
                continue

            # Check for table
            if _is_table_line(stripped):
                if not in_table and current_lines:
                    # Flush non-table content
                    segments.append({
                        "text": "\n".join(current_lines),
                        "section": current_section,
                        "page_start": current_page_start,
                        "page_end": current_page_end,
                        "is_table": False,
                    })
                    current_lines = []
                    current_page_start = page_num
                in_table = True
                current_lines.append(stripped)
                current_page_end = page_num
            else:
                if in_table and current_lines:
                    # Flush table content
                    segments.append({
                        "text": "\n".join(current_lines),
                        "section": current_section,
                        "page_start": current_page_start,
                        "page_end": current_page_end,
                        "is_table": True,
                    })
                    current_lines = []
                    current_page_start = page_num
                    in_table = False
                current_lines.append(stripped)
                current_page_end = page_num

    # Flush final segment
    if current_lines:
        segments.append({
            "text": "\n".join(current_lines),
            "section": current_section,
            "page_start": current_page_start,
            "page_end": current_page_end,
            "is_table": in_table,
        })

    # If no headings were detected, return None to fall back to simple chunking
    headings_found = sum(1 for s in segments if s["section"])
    if headings_found == 0:
        return []

    # Now chunk segments that are too large
    result = []
    for seg in segments:
        text = seg["text"]
        if len(text) <= max_chars:
            if len(text.strip()) >= CHUNK_MIN_CHARS:
                result.append({
                    "text": text,
                    "section": seg["section"],
                    "page_start": seg["page_start"],
                    "page_end": seg["page_end"],
                })
        else:
            # Split large segments by paragraphs
            sub_chunks = chunk_paragraphs(text)
            for sc in sub_chunks:
                result.append({
                    "text": sc,
                    "section": seg["section"],
                    "page_start": seg["page_start"],
                    "page_end": seg["page_end"],
                })

    return result


# ─── PDF Ingestion ─────────────────────────────────────────────────────────────

def ingest_pdf(conn: sqlite3.Connection, source: dict) -> int:
    try:
        import pypdf
    except ImportError:
        print("  ERROR: pypdf not installed. Run: uv add pypdf")
        return 0

    file_path = PROJECT_ROOT / source["path"]
    if not file_path.exists():
        print(f"  ERROR: File not found: {file_path}")
        return 0

    source_dir = CHUNKS_DIR / source["id"]
    clear_source(conn, source["id"])

    domain = source.get("domain", "")
    region = source.get("region", "")
    language = source.get("language", "en")

    count = 0
    now = datetime.now().isoformat()
    reader = pypdf.PdfReader(str(file_path))

    # Extract all pages
    pages: list[tuple[int, str]] = []
    for page_num, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            pages.append((page_num, text))

    # Try semantic chunking first
    semantic_chunks = _semantic_chunk_pages(pages, CHUNK_MAX_CHARS)

    if semantic_chunks:
        for sc in semantic_chunks:
            chunk_id = make_chunk_id(source["id"], count)
            section = sc["section"] or f"Page {sc['page_start']}"
            if sc["page_end"] != sc["page_start"]:
                section += f" (p{sc['page_start']}–{sc['page_end']})"
            save_chunk(conn, {
                "id": chunk_id, "source_id": source["id"], "source_type": "pdf",
                "source_name": source["name"], "file_path": source["path"],
                "section": section, "content": sc["text"], "created_at": now,
                "domain": domain, "region": region, "language": language,
                "page_number": sc["page_start"],
            }, source_dir)
            count += 1
    else:
        # Fallback: original page-buffered 600-char chunking
        buffer, buffer_start = "", None

        for page_num, text in pages:
            if buffer_start is None:
                buffer_start = page_num
            buffer += f"\n\n{text}"

            if len(buffer) >= CHUNK_MAX_CHARS:
                end_page = page_num
                section = f"Page {buffer_start}" + (f"–{end_page}" if end_page != buffer_start else "")
                chunk_id = make_chunk_id(source["id"], count)
                save_chunk(conn, {
                    "id": chunk_id, "source_id": source["id"], "source_type": "pdf",
                    "source_name": source["name"], "file_path": source["path"],
                    "section": section, "content": buffer.strip(), "created_at": now,
                    "domain": domain, "region": region, "language": language,
                    "page_number": buffer_start,
                }, source_dir)
                count += 1
                buffer, buffer_start = "", None

        # Flush remainder
        if buffer.strip():
            section = f"Page {buffer_start}" if buffer_start else "Document"
            chunk_id = make_chunk_id(source["id"], count)
            save_chunk(conn, {
                "id": chunk_id, "source_id": source["id"], "source_type": "pdf",
                "source_name": source["name"], "file_path": source["path"],
                "section": section, "content": buffer.strip(), "created_at": now,
                "domain": domain, "region": region, "language": language,
                "page_number": buffer_start,
            }, source_dir)
            count += 1

    conn.commit()
    return count


# ─── XLSX Ingestion ─────────────────────────────────────────────────────────────

def _extract_country_region(row_text: str, headers: list[str]) -> str:
    """Extract region code from a row if it contains a country column."""
    # Find country column index
    country_idx = -1
    for i, h in enumerate(headers):
        if h.lower().strip() in ("country", "country/region", "destination"):
            country_idx = i
            break
    if country_idx < 0:
        return ""

    # Parse country from the row text (format: "Header: Value | Header: Value")
    for part in row_text.split(" | "):
        if ":" in part:
            key, val = part.split(":", 1)
            key_lower = key.strip().lower()
            if key_lower in ("country", "country/region", "destination"):
                country_name = val.strip().lower()
                # Try exact match first
                if country_name in COUNTRY_REGION_MAP:
                    return COUNTRY_REGION_MAP[country_name]
                # Try partial match
                for name, code in COUNTRY_REGION_MAP.items():
                    if name in country_name or country_name in name:
                        return code
    return ""


def _detect_service_type(sheet_name: str, headers: list[str]) -> str:
    """Detect toll-free/local/DID from sheet name or headers."""
    combined = (sheet_name + " " + " ".join(headers)).lower()
    if "toll-free" in combined or "tollfree" in combined or "toll free" in combined:
        return "toll-free"
    if "did" in combined:
        return "DID"
    if "local" in combined:
        return "local"
    return ""


def ingest_xlsx(conn: sqlite3.Connection, source: dict) -> int:
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl not installed. Run: uv add openpyxl")
        return 0

    file_path = PROJECT_ROOT / source["path"]
    if not file_path.exists():
        print(f"  ERROR: File not found: {file_path}")
        return 0

    source_dir = CHUNKS_DIR / source["id"]
    clear_source(conn, source["id"])

    domain = source.get("domain", "")
    src_region = source.get("region", "")
    language = source.get("language", "en")
    is_rate_table_source = source.get("is_rate_table", False)

    count = 0
    now = datetime.now().isoformat()
    wb = openpyxl.load_workbook(str(file_path), data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows, headers = [], []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(v).strip() if v is not None else "" for v in row]
            if not any(cells):
                continue

            if row_idx == 0:
                headers = cells
            else:
                if headers:
                    row_dict = {h: v for h, v in zip(headers, cells) if h and v}
                    row_text = " | ".join(f"{k}: {v}" for k, v in row_dict.items())
                else:
                    row_text = " | ".join(v for v in cells if v)
                if row_text.strip():
                    rows.append(row_text)

        # Detect rate/pricing tables by header content → use smaller chunk size
        header_lower = {h.lower() for h in headers}
        is_rate_table = bool(header_lower & RATE_TABLE_SIGNALS)
        chunk_size = ROWS_PER_RATE_TABLE_CHUNK if is_rate_table else ROWS_PER_XLSX_CHUNK

        # Detect service type from sheet/headers
        service_type = _detect_service_type(sheet_name, headers)

        # Check if we have a country column for per-row region extraction
        has_country_col = is_rate_table_source and any(
            h.lower().strip() in ("country", "country/region", "destination")
            for h in headers
        )

        # Group rows into chunks by both row count AND character count
        batches = []
        current_batch: list[str] = []
        current_chars = 0
        for row in rows:
            row_len = len(row) + 1  # +1 for newline
            if current_batch and (
                len(current_batch) >= chunk_size
                or current_chars + row_len > CHUNK_MAX_CHARS
            ):
                batches.append(current_batch)
                current_batch = [row]
                current_chars = row_len
            else:
                current_batch.append(row)
                current_chars += row_len
        if current_batch:
            batches.append(current_batch)

        for batch in batches:
            content = f"[Sheet: {sheet_name}]\n" + "\n".join(batch)
            if len(content.strip()) < CHUNK_MIN_CHARS:
                continue

            # Determine region for this chunk
            chunk_region = src_region
            if has_country_col:
                # Extract region from the first row's country field
                regions_in_batch = set()
                for row_text in batch:
                    r = _extract_country_region(row_text, headers)
                    if r:
                        regions_in_batch.add(r)
                if regions_in_batch:
                    chunk_region = "/".join(sorted(regions_in_batch))
                    if service_type:
                        chunk_region += f"/{service_type}"

            chunk_id = make_chunk_id(source["id"], count)
            save_chunk(conn, {
                "id": chunk_id, "source_id": source["id"], "source_type": "xlsx",
                "source_name": source["name"], "file_path": source["path"],
                "section": sheet_name, "content": content, "created_at": now,
                "domain": domain, "region": chunk_region, "language": language,
            }, source_dir)
            count += 1

    conn.commit()
    return count


# ─── Web Ingestion ─────────────────────────────────────────────────────────────

def fetch_page_text(url: str, session) -> tuple[str, str]:
    """Return (page_title, cleaned_text) for a URL."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise ImportError("beautifulsoup4 not installed. Run: uv add beautifulsoup4")

    headers = {"User-Agent": "Mozilla/5.0 (compatible; KBBuilder/1.0)"}
    resp = session.get(url, headers=headers, timeout=15)
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "html.parser")
    for tag in soup(["nav", "footer", "script", "style", "header", "aside"]):
        tag.decompose()

    title = soup.title.string.strip() if soup.title else url
    body = soup.find("main") or soup.find("article") or soup.body or soup

    lines = []
    for el in body.find_all(["h1", "h2", "h3", "h4", "p", "li", "td", "th"]):
        text = el.get_text(separator=" ", strip=True)
        if len(text) > 20:
            prefix = "\n\n## " if el.name in ("h1", "h2", "h3") else ""
            lines.append(f"{prefix}{text}")

    return title, "\n".join(lines)


def get_same_domain_links(url: str, base_url: str, session) -> list[str]:
    try:
        from bs4 import BeautifulSoup
        from urllib.parse import urljoin, urlparse
    except ImportError:
        return []

    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; KBBuilder/1.0)"}
        resp = session.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(resp.text, "html.parser")
        base_parsed = urlparse(base_url)
        links = []
        for a in soup.find_all("a", href=True):
            href = urljoin(url, a["href"])
            parsed = urlparse(href)
            if parsed.netloc == base_parsed.netloc and parsed.scheme in ("http", "https"):
                clean = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
                if clean not in links:
                    links.append(clean)
        return links
    except Exception:
        return []


def ingest_web(conn: sqlite3.Connection, source: dict) -> int:
    import requests

    base_url = source["url"]
    max_pages = source.get("max_pages", 20)
    crawl_depth = source.get("crawl_depth", 1)

    domain = source.get("domain", "")
    region = source.get("region", "")
    language = source.get("language", "en")

    source_dir = CHUNKS_DIR / source["id"]
    clear_source(conn, source["id"])

    count = 0
    now = datetime.now().isoformat()
    session = requests.Session()

    to_visit = [base_url]
    visited: set[str] = set()

    while to_visit and len(visited) < max_pages:
        url = to_visit.pop(0)
        if url in visited:
            continue
        visited.add(url)

        print(f"    Fetching: {url}")
        try:
            title, text = fetch_page_text(url, session)
            time.sleep(0.5)
        except Exception as e:
            print(f"    SKIP ({e})")
            continue

        if not text.strip():
            continue

        # Split by section headings
        sections = re.split(r"\n\n## ", text)
        for sec_text in sections:
            if not sec_text.strip():
                continue
            lines = sec_text.strip().splitlines()
            section_title = lines[0].replace("## ", "").strip()[:80] if lines else title

            for sub_text in chunk_paragraphs(sec_text):
                if len(sub_text.strip()) < CHUNK_MIN_CHARS:
                    continue
                chunk_id = make_chunk_id(source["id"], count)
                save_chunk(conn, {
                    "id": chunk_id, "source_id": source["id"], "source_type": "web",
                    "source_name": source["name"], "source_url": url,
                    "section": section_title, "content": sub_text.strip(), "created_at": now,
                    "domain": domain, "region": region, "language": language,
                }, source_dir)
                count += 1

        if crawl_depth > 1 and len(visited) < max_pages:
            for link in get_same_domain_links(url, base_url, session)[:10]:
                if link not in visited:
                    to_visit.append(link)

    conn.commit()
    print(f"    Visited {len(visited)} pages")
    return count


# ─── Main ──────────────────────────────────────────────────────────────────────

def load_sources() -> dict:
    if not SOURCES_FILE.exists():
        print(f"ERROR: sources.json not found at {SOURCES_FILE}")
        sys.exit(1)
    return json.loads(SOURCES_FILE.read_text(encoding="utf-8"))


def main():
    parser = argparse.ArgumentParser(description="Build knowledge base from web and file sources")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--all", action="store_true", help="Ingest all sources")
    group.add_argument("--source", choices=["web", "files"], help="Ingest by type")
    group.add_argument("--url", help="Ingest a single URL")
    group.add_argument("--file", help="Ingest a single local file path")
    args = parser.parse_args()

    db_path = KB_DIR / "kb.db"
    conn = init_db(db_path)
    sources = load_sources()
    total = 0

    # ── File sources ──
    if args.all or args.source == "files" or args.file:
        file_sources = sources.get("files", [])
        if args.file:
            matched = [s for s in file_sources if s["path"] == args.file]
            if not matched:
                ext = Path(args.file).suffix.lower().lstrip(".")
                matched = [{"id": "adhoc", "path": args.file,
                            "name": Path(args.file).stem, "type": ext}]
            file_sources = matched

        for src in file_sources:
            print(f"\n[KB Ingest] {src['source_type'] if 'source_type' in src else src['type'].upper()}: {src['name']}")
            ftype = src.get("type", "")
            if ftype == "pdf":
                n = ingest_pdf(conn, src)
            elif ftype == "xlsx":
                n = ingest_xlsx(conn, src)
            else:
                print(f"  SKIP: unsupported type '{ftype}'")
                n = 0
            print(f"  → {n} chunks")
            total += n

    # ── Web sources ──
    if args.all or args.source == "web" or args.url:
        web_sources = sources.get("web", [])
        if args.url:
            matched = [s for s in web_sources if s["url"] == args.url]
            if not matched:
                matched = [{"id": "adhoc_web", "url": args.url,
                            "name": args.url, "max_pages": 10, "crawl_depth": 1}]
            web_sources = matched

        for src in web_sources:
            print(f"\n[KB Ingest] WEB: {src['name']} ({src['url']})")
            n = ingest_web(conn, src)
            print(f"  → {n} chunks")
            total += n

    # Update sources metadata (include domain/region)
    meta = {}
    for s in sources.get("files", []):
        meta[s["id"]] = {"name": s["name"], "type": s["type"],
                         "domain": s.get("domain", ""),
                         "region": s.get("region", ""),
                         "updated_at": datetime.now().isoformat()}
    for s in sources.get("web", []):
        meta[s["id"]] = {"name": s["name"], "url": s["url"],
                         "domain": s.get("domain", ""),
                         "region": s.get("region", ""),
                         "updated_at": datetime.now().isoformat()}
    (KB_DIR / "sources_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    conn.close()
    print(f"\n[KB Ingest] Done. Total chunks written: {total}")
    print(f"[KB Ingest] Database: {db_path}")


if __name__ == "__main__":
    main()
